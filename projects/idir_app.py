import tkinter as tk
from tkinter import simpledialog, filedialog, messagebox, ttk
import json
import os
import sys
import tempfile
import shutil
from PIL import Image, ImageTk  # Add this import at the top

class IddirApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Iddir Management")
        self.geometry("700x420")  # Restore to a smaller, previous size
        self.protocol("WM_DELETE_WINDOW", self._on_main_close)  # Confirm on exit

        self.language = "en"
        self.pin_file = "iddir_pin.json"
        self.texts = {
            "en": {
                "home_title": "Iddir Management System",
                "create": "Create a New Iddir File",
                "open": "Open Existing Iddir",
                "language": "Language",
                "iddir_name": "Iddir Name:",
                "admin_name": "Admin Name:",
                "next": "Next",
                "about": "Iddir is a traditional Ethiopian community association that provides mutual aid during times of crisis, especially funerals. Members contribute money regularly to a shared fund, which helps cover burial costs and offers support to grieving families. Over time, Iddirs have expanded to assist with illness, disasters, and financial hardship.",
                "enter_pin": "Enter PIN:",
                "confirm": "Confirm",
                "set_pin": "Set a new PIN:",
                "confirm_pin": "Confirm PIN:",
                "pin_success": "PIN set successfully.",
                "pin_error": "PINs do not match.",
                "pin_required": "PIN required to proceed.",
                "invalid_pin": "Incorrect PIN.",
                "save_as": "Save as:",
                "add_row": "Add Row",
                "delete_row": "Delete Row",
                "delete_row_warning": "Please select a valid member row to delete.",
                "delete_row_confirm": "Do you want to remove {member_name}?",
                "this_member": "this member",
                "monthly_payment": "Monthly Payment:",
                "penalty_amount": "Penalty Amount:"
            },
            "am": {
                "home_title": "የኢዲር አስተዳደር ሲስተም",
                "create": "አዲስ ኢዲር ፍጠር",
                "open": "አሁን ያለውን ኢዲር ክፈት",
                "language": "ቋንቋ",
                "iddir_name": "የኢዲሩ ስም:",
                "admin_name": "የአስተዳደሩ ስም:",
                "next": "ቀጣይ",
                "about": "ኢዲር በኢትዮጵያ የተለመደ ማህበራዊ ድጋፍ ድርጅት ሲሆን በተለይ በቀብር ስዕለት ጊዜ እርዳታን ይሰጣል። አባላቱ በመደበኛነት ገንዘብ ይደርሳሉ፣ ይህም የተአምራትን ክፍያ ለመከፈልና ለቤተሰቡ ድጋፍ ይረዳቸዋል።",
                "enter_pin": "PIN ያስገቡ:",
                "confirm": "አረጋግጥ",
                "set_pin": "አዲስ PIN ያስገቡ:",
                "confirm_pin": "PIN ድጋሚ ያረጋግጡ:",
                "pin_success": "PIN በተሳካ ሁኔታ ተቀመጠ።",
                "pin_error": "PIN አይዛመድም።",
                "pin_required": "ለመቀጠል PIN ያስፈልጋል።",
                "invalid_pin": "PIN ትክክል አይደለም።",
                "save_as": "እንደ ሌላ ያስቀምጡ:",
                "add_row": "ተጨማሪ ረዳት",
                "delete_row": "ረዳት ሰርዝ",
                "delete_row_warning": "እባክዎን የሚሰረዝ ተገቢ ረዳት ይምረጡ።",
                "delete_row_confirm": "{member_name} ረዳት ማስወገድ ይፈልጋሉ?",
                "this_member": "ይህ ረዳት",
                "monthly_payment": "ወርሃዊ ክፍያ:",
                "penalty_amount": "የግድያ ክፍያ:"
            }
        }  # Closing brace added here
        # Add penalty-related attributes
        self.monthly_payment = 0.0
        self.penalty_amount = 0.0
        self._build_home()

        # Add Help/About and Change PIN to main window
        self.menu_bar = tk.Menu(self)
        self.config(menu=self.menu_bar)
        help_menu = tk.Menu(self.menu_bar, tearoff=0)
        help_menu.add_command(label="Help", command=self._show_help_dialog)
        help_menu.add_command(label="About", command=self._show_about_dialog)
        self.menu_bar.add_cascade(label="Help", menu=help_menu)
        account_menu = tk.Menu(self.menu_bar, tearoff=0)
        account_menu.add_command(label="Change PIN", command=self._change_pin_dialog)
        self.menu_bar.add_cascade(label="Account", menu=account_menu)

        self._bind_shortcuts()  # Only call this once here

        self.style = ttk.Style()
        self._setup_modern_style()
        # Uncomment the next line to always show the dashboard/tabs on startup for testing:
        # self._open_excel_gui()

    def _setup_modern_style(self):
        # Modernize ttk widgets
        self.style.theme_use('clam')
        self.style.configure('TButton', font=('Segoe UI', 11), padding=8, relief='flat', borderwidth=0)
        self.style.map('TButton',
            background=[('active', '#005fa3'), ('!active', '#0077b6')],
            foreground=[('active', 'white'), ('!active', 'white')]
        )
        self.style.configure('TNotebook.Tab', font=('Segoe UI', 11, 'bold'), padding=[12, 6])
        self.style.configure('TLabel', font=('Segoe UI', 11))
        self.style.configure('TEntry', font=('Segoe UI', 11), padding=6)
        self.style.configure('TFrame', background='#f7faff')
        # For rounded corners, use a modern theme or custom widget if available

    def _translate(self, key):
        return self.texts[self.language][key]

    def _build_home(self):
        navy = "#001f3f"
        dashboard_bg = navy

        for w in self.winfo_children(): w.destroy()
        self.geometry("700x420")  # Restore to a smaller, previous size
        self.configure(bg=dashboard_bg)

        # --- Create a canvas to allow background image with transparency ---
        canvas = tk.Canvas(self, bg=dashboard_bg, highlightthickness=0, bd=0)
        canvas.place(x=0, y=0, relwidth=1, relheight=1)

        # Load and place the transparent PNG as background (use 'th (1)p.png')
        try:
            bg_img = Image.open("th (1)p.png").convert("RGBA")
            bg_img = bg_img.resize((500, 500), Image.ANTIALIAS)
            self.bg_img_tk = ImageTk.PhotoImage(bg_img)
            canvas.create_image(500, 300, image=self.bg_img_tk, anchor="center")
        except Exception:
            pass

        main_frame = tk.Frame(canvas, bg=dashboard_bg)
        canvas.create_window(0, 0, anchor="nw", window=main_frame, width=700, height=420)

        # --- Consistent, modern fonts and larger sizes ---
        try:
            home_title_font = ("Book Antiqua", 28, 'bold')
            about_font = ("Cascadia Mono", 18)
            button_font = ("Segoe UI", 15, "bold")
            lang_button_font = ("Algerian", 16, 'bold')
        except:
            home_title_font = (None, 28, 'bold')
            about_font = (None, 18)
            button_font = (None, 15, "bold")
            lang_button_font = (None, 16, 'bold')

        # Fix: Use pack for the title so it always appears fully
        title_label = tk.Label(
            main_frame,
            text=self._translate("home_title"),
            bg=dashboard_bg,
            fg="white",
            font=("Book Antiqua", 22, 'bold'),
            pady=10
        )
        title_label.pack(pady=(10, 0), anchor="n", fill="x")

        # Intro note about Iddir (top, centered, larger font, more padding)
        about_frame = tk.Frame(main_frame, bg=dashboard_bg)
        about_frame.pack(fill="x", pady=(40, 0))
        tk.Message(
            about_frame,
            text=self._translate("about"),
            bg=dashboard_bg,
            fg="white",
            width=600,
            font=("Cascadia Mono", 14),
            justify="center",
            padx=10,
            pady=10
        ).pack(pady=(0, 10), anchor="center")

        # Options: Create and Open buttons, centered under intro note, improved style
        btns_frame = tk.Frame(main_frame, bg=dashboard_bg)
        btns_frame.pack(pady=(18, 0))

        # Smaller font and button size for both buttons
        small_btn_font = ("Segoe UI", 10, "bold")
        btn_width = 18
        btn_height = 1

        btn_create = tk.Button(
            btns_frame,
            text="🆕  " + self._translate("create"),
            width=btn_width,
            height=btn_height,
            bg="#0077b6",
            fg="white",
            font=small_btn_font,
            bd=0,
            relief='flat',
            cursor="hand2",
            activebackground="#005fa3",
            activeforeground="white",
            padx=6, pady=2,
            command=self._open_create
        )
        btn_create.pack(pady=6, padx=4, fill="x")

        btn_open = tk.Button(
            btns_frame,
            text="📂  " + self._translate("open"),
            width=btn_width,
            height=btn_height,
            bg="#00b894",
            fg="white",
            font=small_btn_font,
            bd=0,
            relief='flat',
            cursor="hand2",
            activebackground="#00916e",
            activeforeground="white",
            padx=6, pady=2,
            command=self._open_existing
        )
        btn_open.pack(pady=6, padx=4, fill="x")

        # Language switch button (bottom left, more compact)
        tk.Button(
            main_frame,
            text=self._translate("language"),
            bg="#003366",
            fg="white",
            font=("Algerian", 12, 'bold'),
            padx=8, pady=4,
            bd=0,
            relief='flat',
            cursor="hand2",
            activebackground="#005fa3",
            command=self._switch_language_dialog_inplace
        ).place(relx=0.02, rely=0.97, anchor="sw")

    def _save_as(self):
        # Prompt user to select a file path before opening the table GUI
        self._show_progress("Saving file...")
        # --- CHANGED: Save to current file if already opened, else ask for path ---
        file_path = getattr(self, "selected_save_path", None)
        if not file_path:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".iddir",
                filetypes=[("Iddir files", "*.iddir"), ("All files", "*.*")],
                title="Save Iddir Data As"
            )
            if not file_path:
                self._hide_progress()
                return  # User cancelled
            self.selected_save_path = file_path  # Store for later use if needed

        # Update penalty amount from the penalty entry (if present)
        try:
            penalty_val = float(self._excel_widgets['penalty_var'].get())
            self.penalty_amount = penalty_val
        except Exception:
            pass  # Ignore if not a valid float

        # Save the current table data to the selected file (do not open a new iddir)
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            # Save penalty as first row (updated)
            ws.append(["Penalty", self.penalty_amount])
            # Save headers
            columns = self._excel_widgets['columns']
            ws.append(columns)
            # Save data rows (excluding the last 'Total' row)
            data = self._excel_widgets['sheet'].get_sheet_data()
            for row in data[:-1]:
                ws.append(row)
            wb.save(file_path)
            messagebox.showinfo("Success", "Iddir file saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file: {e}")
        self._hide_progress()

    def load_from_iddir(self, iddir_path):
        self._show_progress("Loading file...")
        try:
            """
            Loads data from a .iddir file by copying it to a temp .xlsx file and using openpyxl.
            Returns: (penalty, columns, data) or raises Exception.
            """
            import openpyxl
            temp_dir = tempfile.mkdtemp()
            temp_xlsx = os.path.join(temp_dir, "temp_iddir.xlsx")
            try:
                shutil.copy2(iddir_path, temp_xlsx)
                wb = openpyxl.load_workbook(temp_xlsx)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows or len(rows) < 3:
                    raise Exception("File is missing required data.")
                penalty_row = rows[0]
                if penalty_row[0] != "Penalty":
                    raise Exception("First row must be Penalty.")
                penalty = penalty_row[1]
                columns = list(rows[1])
                data = [list(r) for r in rows[2:]]
                return penalty, columns, data
            finally:
                try:
                    shutil.rmtree(temp_dir)
                except Exception:
                    pass
        finally:
            self._hide_progress()

    def _open_existing(self):
        from tkinter import simpledialog
        file_path = filedialog.askopenfilename(
            defaultextension=".iddir",
            filetypes=[("Iddir files", "*.iddir"), ("All files", "*.*")],
            title="Open Iddir File"
        )
        if not file_path:
            return

        # --- Automatically look for iddir_pin.json in the same folder as the .iddir file ---
        iddir_dir = os.path.dirname(file_path)
        pin_file_candidate = os.path.join(iddir_dir, "iddir_pin.json")
        if os.path.exists(pin_file_candidate):
            self.pin_file = pin_file_candidate
        elif os.path.exists(self.pin_file):
            # Use default location if exists
            pass
        else:
            messagebox.showerror("Error", "PIN file missing. Please place iddir_pin.json in the same folder as your .iddir file or in the app folder.")
            return

        with open(self.pin_file) as f:
            iddir_data = json.load(f)
        pin_expected = iddir_data.get("pin")
        self.monthly_payment = iddir_data.get("monthly_payment", 0.0)
        self.penalty_amount = iddir_data.get("penalty_amount", 0.0)

        # Prompt for PIN
        pin = simpledialog.askstring("PIN", self._translate("enter_pin") + "\n(Forgot PIN? Contact admin.)", show='*')
        if not pin:
            messagebox.showerror("Error", self._translate("pin_required"))
            return
        if pin != pin_expected:
            messagebox.showerror("Error", self._translate("invalid_pin"))
            return

        iddir_data_for_gui = {
            "monthly_payment": self.monthly_payment,
            "penalty_amount": self.penalty_amount,
            "calendar": iddir_data.get("calendar", "gregorian"),
        }

        try:
            penalty, columns, data = self.load_from_iddir(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file: {e}")
            return
        self._open_excel_gui(data=data, iddir_data=iddir_data_for_gui)
        if hasattr(self, '_excel_widgets') and 'penalty_var' in self._excel_widgets:
            self._excel_widgets['penalty_var'].set(str(penalty))

    def _open_excel_gui(self, data=None, iddir_data=None):
        try:
            import tksheet
        except ImportError:
            messagebox.showerror("Missing Dependency", "Please install tksheet: pip install tksheet")
            return
        if iddir_data is not None:
            # Use data passed from _open_existing
            self.monthly_payment = iddir_data.get("monthly_payment", 0.0)
            self.penalty_amount = iddir_data.get("penalty_amount", 0.0)
            calendar_type = iddir_data.get("calendar", "gregorian")
        else:
            # Creating new file: load from iddir_pin.json
            if not os.path.exists(self.pin_file):
                messagebox.showerror("Error", "Please confirm and save Iddir info first.")
                return
            with open(self.pin_file) as f:
                iddir_data = json.load(f)
            self.monthly_payment = iddir_data.get("monthly_payment", 0.0)
            self.penalty_amount = iddir_data.get("penalty_amount", 0.0)
            calendar_type = self.cal_type.get() if hasattr(self, 'cal_type') else iddir_data.get("calendar", "gregorian")
        gregorian_months = [
            "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"
        ]
        ethiopian_months = [
            "September", "October", "November", "December", "January", "February", "March", "April", "May", "June", "July", "August"
        ]
        amharic_gregorian = [
            "ጃንዩወሪ", "ፌብሩወሪ", "ማርች", "ኤፕሪል", "ሜይ", "ጁን", "ጁላይ", "ኦገስት", "ሴፕቴምበር", "ኦክቶበር", "ኖቬምበር", "ዲሴምበር"
        ]
        amharic_ethiopian = [
            "መስከረም", "ጥቅምት", "ኅዳር", "ታኅሣሥ", "ጥር", "የካቲት", "መጋቢት", "ሚያዝያ", "ግንቦት", "ሰኔ", "ሀምሌ", "ነሐሴ"
        ]
        if self.language == 'am':
            months = amharic_gregorian if calendar_type == "gregorian" else amharic_ethiopian
            total_label = "ድምር"
        else:
            months = gregorian_months if calendar_type == "gregorian" else ethiopian_months
            total_label = "Total"
        columns = ["Members"] + months
        # Prepare data with 'Total' row as the last row
        if data is None:
            data = [["" for _ in columns] for _ in range(10)]
        else:
            # Always show at least 10 blank rows if no data
            filtered_data = []
            for row in data:
                if row and row[0] not in ("Total", "ድምር"):
                    filtered_data.append(list(row) + [""] * (len(columns) - len(row)))
            if not filtered_data:
                data = [["" for _ in columns] for _ in range(10)]
            else:
                data = filtered_data
        data.append([total_label] + ["" for _ in months])

        # --- Create the main window for the tabs ---
        win = tk.Toplevel(self)
        win.title("Iddir Members Table")
        win.geometry("1100x600")  # Ensure window is always large enough
        win.minsize(1100, 600)    # Prevent resizing below this size
        win.configure(bg="#f7faff")
        win.grab_set()
        self._excel_widgets = {}

        # --- Gradient background for the main window ---
        def draw_gradient(canvas, width, height, color1, color2):
            import colorsys
            r1, g1, b1 = win.winfo_rgb(color1)
            r2, g2, b2 = win.winfo_rgb(color2)
            r1, g1, b1 = r1//256, g1//256, b1//256
            r2, g2, b2 = r2//256, g2//256, b2//256
            steps = height
            for i in range(steps):
                r = int(r1 + (r2 - r1) * i / steps)
                g = int(g1 + (g2 - g1) * i / steps)
                b = int(b1 + (b2 - b1) * i / steps)
                color = f"#{r:02x}{g:02x}{b:02x}"
                canvas.create_line(0, i, width, i, fill=color)

        gradient_canvas = tk.Canvas(win, width=1100, height=600, highlightthickness=0)
        gradient_canvas.place(x=0, y=0, relwidth=1, relheight=1)
        draw_gradient(gradient_canvas, 1100, 600, "#b6e0fe", "#f7faff")

        notebook = ttk.Notebook(win)
        notebook.pack(fill="both", expand=True)

        def create_tab_with_gradient(parent, width=1100, height=600, color1="#e3f0ff", color2="#f7faff"):
            frame = tk.Frame(parent, bg=color2)
            canvas = tk.Canvas(frame, width=width, height=height, highlightthickness=0)
            canvas.place(x=0, y=0, relwidth=1, relheight=1)
            draw_gradient(canvas, width, height, color1, color2)
            content = tk.Frame(frame, bg="", highlightthickness=0)
            content.place(x=0, y=0, relwidth=1, relheight=1)
            return frame, content

        table_tab, table_content = create_tab_with_gradient(notebook, 1100, 600, "#e3f0ff", "#f7faff")
        notebook.add(table_tab, text="📋 Table")
        report_tab, report_content = create_tab_with_gradient(notebook, 1100, 600, "#f7faff", "#e3f0ff")
        notebook.add(report_tab, text="📊 Report")
        penalty_tab, penalty_content = create_tab_with_gradient(notebook, 1100, 600, "#e3f0ff", "#f7faff")
        notebook.add(penalty_tab, text="⚠️ Penalty List")

        # --- Penalty/Monthly info at top of Table tab ---
        info_frame = tk.Frame(table_content, bg="#f7faff")
        info_frame.pack(side="top", fill="x", padx=10, pady=(10, 0))
        tk.Label(info_frame, text=f"Monthly Payment: {self.monthly_payment}", bg="#f7faff", fg="#003366", font=("Algerian", 12, "bold")).pack(side="left", padx=(0, 20))
        tk.Label(info_frame, text=f"Penalty: {self.penalty_amount}", bg="#f7faff", fg="#d90429", font=("Algerian", 12, "bold")).pack(side="left", padx=(0, 20))

        # --- Year Selection Button ---
        # (Removed year button and functionality)

        # --- Penalty Entry Box (attractive, packed) ---
        penalty_frame = tk.Frame(table_content, bg="#f7faff")
        penalty_frame.pack(side="top", fill="x", padx=10, pady=(10, 0))
        penalty_var = tk.StringVar()
        penalty_entry = tk.Entry(penalty_frame, textvariable=penalty_var, width=12, font=(None, 12, "bold"),
                                 bg="#fff0f3", fg="#d90429", highlightthickness=2, highlightbackground="#d90429", relief="solid")
        penalty_entry.pack(side="left")
        penalty_entry.insert(0, "0.0")
        self._excel_widgets['penalty_var'] = penalty_var

        # --- Table (tksheet) ---
        sheet = tksheet.Sheet(
            table_content,
            data=data,
            headers=columns,
            width=980,
            height=440,
            show_x_scrollbar=True,
            show_y_scrollbar=True,
            show_index=False,
            header_height=36,
            outline_thickness=1
        )
        sheet.pack(side="top", fill="both", expand=True, padx=10, pady=10)
        # Enable all editing features for the table
        sheet.enable_bindings((
            "single_select", "row_select", "column_select", "drag_select",
            "column_drag_and_drop", "row_drag_and_drop",
            "column_resize", "row_resize",
            "edit_cell", "arrowkeys", "right_click_popup_menu",
            "rc_select", "copy", "cut", "paste", "delete", "undo", "redo",
            "cell_select"
        ))
        self._excel_widgets['win'] = win
        self._excel_widgets['sheet'] = sheet
        self._excel_widgets['columns'] = columns
        self._excel_widgets['calendar_type'] = calendar_type
        self._excel_widgets['data'] = data

        # --- Payment enforcement logic ---
        def enforce_payment(event=None):
            # Only allow monthly_payment or blank in month cells (except Total row)
            sheet_data = sheet.get_sheet_data()
            num_rows = sheet.get_total_rows()
            num_cols = len(columns)
            for row in range(num_rows - 1):  # skip Total row
                for col in range(1, num_cols):
                    val = sheet.get_cell_data(row, col)
                    if val == "":
                        continue
                    try:
                        v = float(val)
                        if v != float(self.monthly_payment):
                            sheet.set_cell_data(row, col, "")
                    except Exception:
                        sheet.set_cell_data(row, col, "")
            sheet.refresh()
        sheet.bind("<<SheetCellEdited>>", enforce_payment, add='+')

        # --- Prevent editing the 'Total' row except for the label cell ---
        def block_total_row_edit(event):
            edited = sheet.get_edit_cell()
            if not edited:
                return
            row, col = edited
            if row == sheet.get_total_rows() - 1 and col != 0:
                sheet.set_cell_data(row, col, "")
        sheet.bind("<<SheetCellEdited>>", block_total_row_edit, add='+')

        # --- Live column sum logic for months, ignoring 'PENALTY' ---
        def update_column_totals(event=None):
            sheet_data = sheet.get_sheet_data()
            num_rows = sheet.get_total_rows()
            num_cols = len(columns)
            for col in range(1, num_cols):
                col_sum = 0.0
                for row in range(num_rows - 1):
                    val = sheet.get_cell_data(row, col)
                    if isinstance(val, str) and val.strip().upper() == 'PENALTY':
                        continue
                    try:
                        col_sum += float(val)
                    except (ValueError, TypeError):
                        continue
                sheet.set_cell_data(num_rows - 1, col, str(col_sum) if col_sum != 0 else "")
            sheet.set_cell_data(num_rows - 1, 0, total_label)
        sheet.bind("<<SheetCellEdited>>", update_column_totals, add='+')
        update_column_totals()

        # --- Month navigation, autofill, auto sum, add/delete row, etc. ---
        btns_frame = tk.Frame(table_content, bg="#f7faff")
        btns_frame.pack(side="top", fill="x", padx=10, pady=4)

        def switch_month_language_inplace():
            self.language = 'am' if self.language == 'en' else 'en'
            calendar_type = self._excel_widgets['calendar_type']
            if self.language == 'am':
                months = amharic_gregorian if calendar_type == "gregorian" else amharic_ethiopian
                total_label = "ድምር"
            else:
                months = gregorian_months if calendar_type == "gregorian" else ethiopian_months
                total_label = "Total"
            columns = ["Members"] + months
            self._excel_widgets['columns'] = columns
            # Remove any previous total row, then append new one
            sheet_data = self._excel_widgets['sheet'].get_sheet_data()
            filtered = [row for row in sheet_data if row[0] != "Total" and row[0] != "ድምር"]
            # Pad all rows to the new column count
            for row in filtered:
                while len(row) < len(columns):
                    row.append("")
                while len(row) > len(columns):
                    row.pop()
            filtered.append([total_label] + ["" for _ in months])
            self._excel_widgets['sheet'].headers(columns)
            self._excel_widgets['sheet'].set_sheet_data(filtered, reset_col_positions=True, reset_row_positions=True)
            self._excel_widgets['sheet'].refresh()
            self._excel_widgets['btn_lang'].config(text=self._translate("language"))
            self._excel_widgets['btn_add_row'].config(text=self._translate("add_row"))
            self._excel_widgets['btn_delete_row'].config(text=self._translate("delete_row"))

        btn_lang = tk.Button(btns_frame, text="🌐 " + self._translate("language"), bg="#003366", fg="white", font=("Segoe UI", 11), command=switch_month_language_inplace, bd=0, relief='flat', cursor="hand2", activebackground="#005fa3")
        btn_lang.pack(side="left", padx=5)
        self._excel_widgets['btn_lang'] = btn_lang

        def add_row():
            # Insert above the last row (the total row)
            self._excel_widgets['sheet'].insert_row(["" for _ in self._excel_widgets['columns']], idx=self._excel_widgets['sheet'].get_total_rows()-1)
        btn_add_row = tk.Button(btns_frame, text="➕ " + self._translate("add_row"), bg="#0077b6", fg="white", font=("Segoe UI", 11, "bold"), command=add_row, bd=0, relief='flat', cursor="hand2", activebackground="#005fa3")
        btn_add_row.pack(side="left", padx=5)
        self._excel_widgets['btn_add_row'] = btn_add_row

        def delete_row():
            selected = self._excel_widgets['sheet'].get_selected_rows()
            idx = None
            if selected:
                idx = next(iter(selected))
            else:
                selected_cells = self._excel_widgets['sheet'].get_selected_cells()
                if selected_cells:
                    idx = selected_cells[0][0]  # (row, col)
            if idx is None or idx == self._excel_widgets['sheet'].get_total_rows() - 1:
                messagebox.showwarning(self._translate("delete_row"), self._translate("delete_row_warning"))
                return
            member_name = self._excel_widgets['sheet'].get_cell_data(idx, 0)
            if not member_name:
                member_name = self._translate("this_member")
            confirm_msg = self._translate("delete_row_confirm").format(member_name=member_name)
            if messagebox.askyesno(self._translate("delete_row"), confirm_msg):
                self._excel_widgets['sheet'].delete_row(idx)

        btn_delete_row = tk.Button(btns_frame, text="🗑️ " + self._translate("delete_row"), bg="#d90429", fg="white", font=("Segoe UI", 11, "bold"), command=delete_row, bd=0, relief='flat', cursor="hand2", activebackground="#a3001b")
        btn_delete_row.pack(side="left", padx=5)
        self._excel_widgets['btn_delete_row'] = btn_delete_row

        # --- Custom attractive Auto Sum dialog ---
        def show_auto_sum_dialog(total):
            sum_win = tk.Toplevel(win)
            sum_win.title("Auto Sum Result")
            sum_win.configure(bg="#f7faff")
            sum_win.geometry("320x120")
            sum_win.resizable(False, False)
            sum_win.grab_set()
            tk.Label(sum_win, text="Sum of selected cells:", bg="#f7faff", fg="#222", font=(None, 13)).pack(pady=(18, 2))
            tk.Label(sum_win, text=f"{total}", bg="#f7faff", fg="#d90429", font=(None, 22, "bold")).pack(pady=(0, 10))
            tk.Button(sum_win, text="OK", command=sum_win.destroy, bg="#0984e3", fg="white", font=(None, 11, "bold"), width=8).pack(pady=(0, 8))

        def auto_sum_selected():
            selected_cells = sheet.get_selected_cells()
            total = 0.0
            for row, col in selected_cells:
                val = sheet.get_cell_data(row, col)
                if isinstance(val, str) and val.strip().upper() == 'PENALTY':
                    continue
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    continue
            show_auto_sum_dialog(total)

        btn_auto_sum = tk.Button(btns_frame, text="∑ Auto Sum", bg="#00b894", fg="white", font=("Segoe UI", 11, "bold"), command=auto_sum_selected, bd=0, relief='flat', cursor="hand2", activebackground="#00916e")
        btn_auto_sum.pack(side="left", padx=5)
        self._excel_widgets['btn_auto_sum'] = btn_auto_sum

        # --- Autofill selected cells (Excel-style) ---
        def autofill_selected_cells():
            selected = sheet.get_selected_cells()
            if not selected:
                messagebox.showinfo("Autofill", "Please select at least two cells.")
                return
            selected = sorted(selected)
            first_row, first_col = selected[0]
            value = sheet.get_cell_data(first_row, first_col)
            for row, col in selected:
                sheet.set_cell_data(row, col, value)
            sheet.refresh()

        btn_autofill = tk.Button(btns_frame, text="📝 Autofill", bg="#fdcb6e", fg="black", font=("Segoe UI", 11, "bold"), command=autofill_selected_cells, bd=0, relief='flat', cursor="hand2", activebackground="#ffeaa7")
        btn_autofill.pack(side="left", padx=5)

        btn_save_excel = tk.Button(
            btns_frame,
            text="💾 Save",
            bg="#0984e3",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            bd=0,
            relief='flat',
            cursor="hand2",
            activebackground="#005fa3",
            command=self._save_as  # Restore save button functionality
        )
        btn_save_excel.pack(side="left", padx=5)
        self._excel_widgets['btn_save_excel'] = btn_save_excel

        # --- Live column sum logic for months, ignoring 'PENALTY' ---
        def update_column_totals(event=None):
            sheet_data = sheet.get_sheet_data()
            num_rows = sheet.get_total_rows()
            num_cols = len(columns)
            # The last row is the 'Total' row
            for col in range(1, num_cols):  # skip 'Members' col
                col_sum =  0.0
                for row in range(num_rows - 1):  # skip last row (Total)
                    val = sheet.get_cell_data(row, col)
                    if isinstance(val, str) and val.strip().upper() == 'PENALTY':
                        continue
                    try:
                        col_sum += float(val)
                    except (ValueError, TypeError):
                        continue
                # Set the total in the last row for this column
                sheet.set_cell_data(num_rows - 1, col, str(col_sum) if col_sum != 0 else "")
            # Always set the label for the first cell of the total row
            sheet.set_cell_data(num_rows - 1, 0, total_label)

        # Bind to cell edit event for live update
        sheet.bind("<<SheetCellEdited>>", update_column_totals)

        # Prevent editing the 'Total' row except for the label cell
        def block_total_row_edit(event):
            edited = sheet.get_edit_cell()
            if not edited:
                return
            row, col = edited
            if row == sheet.get_total_rows() - 1 and col != 0:
                # Revert edit
                sheet.set_cell_data(row, col, "")
        sheet.bind("<<SheetCellEdited>>", block_total_row_edit, add='+')

        # Initial calculation
        update_column_totals()

        # --- Month buttons row (below table) ---
        month_btns_frame = tk.Frame(table_content, bg="#f7faff")
        month_btns_frame.pack(side="top", fill="x", padx=10, pady=(0, 8))
        def open_journal_book(month_idx, month_name):
            # Create a journal book window for the selected month
            jb_win = tk.Toplevel(win)
            jb_win.title(f"Journal Book - {month_name}")
            jb_win.geometry("500x400")
            jb_win.configure(bg="#f7faff")
            jb_win.grab_set()
            tk.Label(jb_win, text=f"Journal Book for {month_name}", bg="#f7faff", fg="#003366", font=(None, 14, "bold")).pack(pady=(18, 8))
            text_area = tk.Text(jb_win, width=60, height=16, font=(None, 11), bg="#fff", fg="#222", wrap="word", borderwidth=2, relief="groove")
            text_area.pack(padx=18, pady=(0, 12), fill="both", expand=True)
            # Load previous journal if exists
            iddir_name = self._excel_widgets['columns'][0] if 'columns' in self._excel_widgets else "iddir"
            journal_dir = os.path.join(os.getcwd(), "journals")
            os.makedirs(journal_dir, exist_ok=True)
            journal_file = os.path.join(journal_dir, f"{month_name}_journal.txt")
            if os.path.exists(journal_file):
                with open(journal_file, "r", encoding="utf-8") as jf:
                    text_area.insert("1.0", jf.read())
            def save_journal():
                with open(journal_file, "w", encoding="utf-8") as jf:
                    jf.write(text_area.get("1.0", "end-1c"))
                messagebox.showinfo("Saved", f"Journal for {month_name} saved.")
            save_btn = tk.Button(jb_win, text="Save", bg="#0984e3", fg="white", font=(None, 11, "bold"), command=save_journal)
            save_btn.pack(pady=(0, 12))
        # Add a button for each month (skip 'Members' col)
        for idx, month in enumerate(columns[1:], 1):
            btn = tk.Button(month_btns_frame, text="📅 " + month, width=12, bg="#00b894", fg="white", font=("Segoe UI", 10, "bold"), command=lambda i=idx, m=month: open_journal_book(i, m), bd=0, relief='flat', cursor="hand2", activebackground="#00916e")
            btn.pack(side="left", padx=3, pady=2)

        # --- Report Tab Content (placeholder) ---
        report_label = tk.Label(report_content, text="Report will be shown here.", bg="#f7faff", fg="#003366", font=(None, 14, "bold"))
        report_label.pack(pady=40)

        # Remove "Description" column, keep only Months, Income, Expense, Balance
        report_columns = ["Months", "Income", "Expense", "Balance"]
        months_list = months
        report_data = []
        for m in months_list:
            report_data.append([m, "", "", ""])

        # Create the tksheet for the report tab (make it larger and editable)
        try:
            import tksheet
        except ImportError:
            messagebox.showerror("Missing Dependency", "Please install tksheet: pip install tksheet")
            return
        report_sheet = tksheet.Sheet(report_content,
                                     data=report_data,
                                     headers=report_columns,
                                     width=800,
                                     height=300,
                                     show_x_scrollbar=True,
                                     show_y_scrollbar=True,
                                     show_index=False,
                                     header_height=36,
                                     outline_thickness=1)
        report_sheet.pack(side="top", fill="x", padx=10, pady=10)
        report_sheet.enable_bindings((
            "single_select", "row_select", "column_select", "drag_select",
            "column_drag_and_drop", "row_drag_and_drop",
            "column_resize", "row_resize",
            "edit_cell", "arrowkeys", "right_click_popup_menu",
            "rc_select", "copy", "cut", "paste", "delete", "undo", "redo",
            "cell_select"
        ))

        def report_edit_validation(event):
            row, col = report_sheet.get_edit_cell()
            if col == 0:
                report_sheet.set_cell_data(row, col, months_list[row])
                return
            if col in [1,2,3]:
                value = report_sheet.get_cell_data(row, col)
                if value == "":
                    return
                try:
                    float(value)
                except ValueError:
                    report_sheet.set_cell_data(row, col, "")
        report_sheet.bind("<<SheetCellEdited>>", report_edit_validation)

        report_instr = tk.Label(
            report_content,
            text="Only numbers are allowed in Income, Expense, and Balance columns.",
            bg="#f7faff", fg="#d90429", font=(None, 10, "italic")
        )
        report_instr.pack(pady=(0, 2))

        # Add note about expenditures
        report_note = tk.Label(
            report_content,
            text="All expenditures are assumed to be 50 less than the total paid monthly amount.",
            bg="#f7faff", fg="#003366", font=(None, 10, "italic")
        )
        report_note.pack(pady=(0, 10))

        def import_income_to_report():
            main_sheet = sheet
            num_rows = main_sheet.get_total_rows()
            for month_idx, month_name in enumerate(months_list):
                col_idx = main_sheet.headers().index(month_name) if month_name in main_sheet.headers() else None
                if col_idx is None:
                    continue
                col_sum = 0.0
                not_paid_count = 0
                for row in range(num_rows - 1):
                    val = main_sheet.get_cell_data(row, col_idx)
                    try:
                        v = float(val)
                        col_sum += v
                        if v == 0:
                            not_paid_count += 1
                    except (ValueError, TypeError):
                        # Treat empty or non-numeric as not paid
                        not_paid_count += 1
                # Add penalty for each person who didn't pay
                penalty_income = not_paid_count * float(self.penalty_amount)
                income_with_penalty = col_sum + penalty_income
                report_sheet.set_cell_data(month_idx, 1, str(income_with_penalty) if income_with_penalty != 0 else "")
                # Expense: income - 50
                expense = income_with_penalty - 50
                report_sheet.set_cell_data(month_idx, 2, str(expense))
                # Balance: income - expense
                balance = income_with_penalty - expense
                report_sheet.set_cell_data(month_idx, 3, str(balance))
            report_sheet.refresh()

        btn_import_income = tk.Button(report_content, text="Import Income", bg="#00b894", fg="white", font=(None, 11, "bold"), command=import_income_to_report)
        btn_import_income.pack(side="bottom", fill="x", padx=10, pady=10)

        # --- Restore Penalty Tab Table ---
        # Remove previous penalty sheet and summary if present
        for widget in penalty_content.winfo_children():
            widget.destroy()

        # Build penalty table columns and data
        penalty_columns = ["Member", "Month", "Amount", "Status"]  # Changed last column to "Status"
        sheet_data = self._excel_widgets['sheet'].get_sheet_data() if 'sheet' in self._excel_widgets else data
        penalty_table_data = []
        penalized_members = set()
        # --- CHANGED: Use actual penalty amount for display ---
        penalty_amount_display = f"{self.penalty_amount} Birr"

        for row in sheet_data:
            if not row or not row[0]:
                continue
            member = row[0]
            if member in ["Total", "ድምር"]:
                continue
            for m_idx, month in enumerate(months, 1):
                val = row[m_idx] if m_idx < len(row) else ""
                try:
                    paid = float(val)
                except Exception:
                    paid = None
                # Penalize if not a valid positive number (empty, text, or zero/negative)
                if paid is None or paid <= 0:
                    penalty_table_data.append([member, month, penalty_amount_display, "Penalized"])
                    penalized_members.add((member, month))

        if not penalty_table_data:
            penalty_table_data.append(["", "", "", ""])

        # Create the tksheet for the penalty tab (table style)
        try:
            import tksheet
        except ImportError:
            messagebox.showerror("Missing Dependency", "Please install tksheet: pip install tksheet")
            return
        penalty_sheet_full = tksheet.Sheet(
            penalty_content,
            data=penalty_table_data,
            headers=penalty_columns,
            width=600,
            height=340,
            show_x_scrollbar=True,
            show_y_scrollbar=True,
            show_index=False,
            header_height=36,
            outline_thickness=1
        )
        penalty_sheet_full.pack(side="top", fill="x", padx=10, pady=(10, 2))
        penalty_sheet_full.enable_bindings((
            "single_select", "row_select", "column_select", "drag_select",
            "column_drag_and_drop", "row_drag_and_drop",
            "column_resize", "row_resize",
            "edit_cell", "arrowkeys", "right_click_popup_menu",
            "rc_select", "copy", "cut", "paste", "delete", "undo", "redo",
            "cell_select"
        ))

        # Show penalized summary at the bottom
        if penalized_members:
            summary_text = "\n".join(
                f"{member} was penalized 50 Birr in month {month}"
                for member, month in penalized_members
            )
        else:
            summary_text = "No one was penalized."
        summary_label = tk.Label(penalty_content, text=summary_text, bg="#f7faff", fg="#d90429", font=(None, 12, "bold"), anchor="w", justify="left")
        summary_label.pack(side="top", fill="x", padx=10, pady=(8, 10))
        self._excel_widgets['penalty_sheet'] = penalty_sheet_full

    def _open_create(self):
        navy = "#001f3f"
        white = "#ffffff"
        from tkinter import ttk
        dlg = tk.Toplevel(self)
        dlg.title(self._translate("create"))
        dlg.geometry("600x550")
        dlg.configure(bg=white)
        dlg.grab_set()

        # Use Bookman Old Style font for create dialog
        try:
            dashboard_font = ("Segoe UI", 18, 'bold')
            button_font = ("Segoe UI", 12, 'bold')
        except:
            dashboard_font = (None, 18, 'bold')
            button_font = (None, 12, 'bold')

        self._create_widgets = {}

        header = tk.Frame(dlg, bg=navy, height=60)
        header.pack(fill='x')
        lbl_create = tk.Label(header, text=self._translate("create"), bg=navy, fg=white,
                font=dashboard_font)
        lbl_create.pack(pady=15)
        self._create_widgets['lbl_create'] = lbl_create

        content = tk.Frame(dlg, bg=white)
        content.pack(fill='both', expand=True, padx=20, pady=10)

        left_col = tk.Frame(content, bg=white)
        left_col.grid(row=0, column=0, sticky='nsew', padx=(0, 10))
        right_col = tk.Frame(content, bg=white)
        right_col.grid(row=0, column=2, sticky='nsew', padx=(10, 0))

        sep = ttk.Separator(content, orient='vertical')
        sep.grid(row=0, column=1, sticky='ns', pady=10)

        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(2, weight=1)
        content.grid_rowconfigure(0, weight=1)

        # Validation functions
        def validate_letters(P):
            return all(c.isalpha() or c.isspace() for c in P)
        def validate_pin(P):
            return P.isdigit() or P == ""
        def validate_float(P):
            try:
                if P == "":
                    return True
                float(P)
                return True
            except:
                return False

        vcmd_letters = (dlg.register(validate_letters), '%P')
        vcmd_pin = (dlg.register(validate_pin), '%P')
        vcmd_float = (dlg.register(validate_float), '%P')

        lbl_iddir_name = tk.Label(left_col, text=self._translate("iddir_name"), bg=white, fg=navy)
        lbl_iddir_name.pack(anchor='w', pady=(0,5))
        self.entry_name = tk.Entry(left_col, width=30, validate='key', validatecommand=vcmd_letters)
        self.entry_name.pack(pady=(0,10), anchor='w')
        lbl_admin_name = tk.Label(left_col, text=self._translate("admin_name"), bg=white, fg=navy)
        lbl_admin_name.pack(anchor='w', pady=(0,5))
        self.entry_admin = tk.Entry(left_col, width=30, validate='key', validatecommand=vcmd_letters)
        self.entry_admin.pack(pady=(0,10), anchor='w')
        lbl_calendar = tk.Label(left_col, text="Calendar Type:", bg=white, fg=navy)
        lbl_calendar.pack(anchor='w', pady=(10,0))
        cal_frame = tk.Frame(left_col, bg=white)
        cal_frame.pack(anchor='w', pady=(0,10))
        self.cal_type = tk.StringVar(value="gregorian")
        def set_calendar_type():
            pass
        rb_eth = tk.Radiobutton(cal_frame, text="Ethiopian", variable=self.cal_type, value="ethiopian",
                    bg="#003366", fg="white", selectcolor="#003366", font=(None, 10), command=set_calendar_type)
        rb_eth.pack(side="left", padx=5)
        rb_greg = tk.Radiobutton(cal_frame, text="Gregorian", variable=self.cal_type, value="gregorian",
                    bg="#003366", fg="white", selectcolor="#003366", font=(None, 10), command=set_calendar_type)
        rb_greg.pack(side="left", padx=5)

        # --- Monthly Payment and Penalty Amount fields ---
        lbl_monthly = tk.Label(left_col, text="Monthly Payment:", bg=white, fg=navy)
        lbl_monthly.pack(anchor='w', pady=(10, 0))
        self.entry_monthly = tk.Entry(left_col, width=20, validate='key', validatecommand=vcmd_float)
        self.entry_monthly.pack(anchor='w', pady=(0, 10))
        lbl_penalty = tk.Label(left_col, text="Penalty Amount:", bg=white, fg=navy)
        lbl_penalty.pack(anchor='w', pady=(0, 0))
        self.entry_penalty = tk.Entry(left_col, width=20, validate='key', validatecommand=vcmd_float)
        self.entry_penalty.pack(anchor='w', pady=(0, 10))

        self._create_widgets.update({
            'lbl_iddir_name': lbl_iddir_name,
            'lbl_admin_name': lbl_admin_name,
            'lbl_calendar': lbl_calendar,
            'rb_eth': rb_eth,
            'rb_greg': rb_greg,
            'lbl_monthly': lbl_monthly,
            'lbl_penalty': lbl_penalty,
            'entry_monthly': self.entry_monthly,
            'entry_penalty': self.entry_penalty
        })

        lbl_set_pin = tk.Label(right_col, text=self._translate("set_pin"), bg=white, fg=navy)
        lbl_set_pin.pack(anchor='w', pady=(0,5))
        self.pin1_entry = tk.Entry(right_col, show='*', width=30, validate='key', validatecommand=vcmd_pin)
        self.pin1_entry.pack(pady=(0,10), anchor='w')
        lbl_confirm_pin = tk.Label(right_col, text=self._translate("confirm_pin"), bg=white, fg=navy)
        lbl_confirm_pin.pack(anchor='w', pady=(0,5))
        self.pin2_entry = tk.Entry(right_col, show='*', width=30, validate='key', validatecommand=vcmd_pin)
        self.pin2_entry.pack(pady=(0,10), anchor='w')

        def save_iddir_info():
            name = self.entry_name.get().strip()
            admin = self.entry_admin.get().strip()
            cal_type = self.cal_type.get()
            pin1 = self.pin1_entry.get()
            pin2 = self.pin2_entry.get()
            monthly = self.entry_monthly.get().strip()
            penalty = self.entry_penalty.get().strip()
            # Validate all fields
            if not name or not admin or not pin1 or not pin2 or not monthly or not penalty:
                messagebox.showerror("Error", self._translate("pin_required"))
                return
            if not validate_letters(name):
                messagebox.showerror("Error", "Iddir Name must contain only letters and spaces.")
                return
            if not validate_letters(admin):
                messagebox.showerror("Error", "Admin Name must contain only letters and spaces.")
                return
            if not validate_pin(pin1) or not validate_pin(pin2):
                messagebox.showerror("Error", "PIN must be digits only.")
                return
            if pin1 != pin2:
                messagebox.showerror("Error", self._translate("pin_error"))
                return
            try:
                monthly_f = float(monthly)
                penalty_f = float(penalty)
            except:
                messagebox.showerror("Error", "Monthly payment and penalty must be numbers.")
                return
            # Save all iddir info (without date)
            with open(self.pin_file, 'w') as f:
                json.dump({
                    "iddir_name": name,
                    "admin": admin,
                    "calendar": cal_type,
                    "pin": pin1,
                    "monthly_payment": monthly_f,
                    "penalty_amount": penalty_f
                }, f)
            self.monthly_payment = monthly_f
            self.penalty_amount = penalty_f
            messagebox.showinfo("Success", self._translate("pin_success"))
            # Do NOT destroy dlg here; keep the create dialog open for Save As
        btn_confirm = tk.Button(right_col, text="✔️ " + self._translate("confirm"), bg="#001f3f", fg="white",
                  font=button_font, command=save_iddir_info, bd=0, relief='flat', cursor="hand2", activebackground="#005fa3")
        btn_confirm.pack(pady=(20,0), anchor='w')
        self._create_widgets.update({
            'lbl_set_pin': lbl_set_pin,
            'lbl_confirm_pin': lbl_confirm_pin,
            'btn_confirm': btn_confirm
        })

        btn_lang = tk.Button(dlg, text="🌐 " + self._translate("language"), bg="#003366", fg="white",
                font=button_font, command=lambda: self._switch_language_dialog_inplace(), bd=0, relief='flat', cursor="hand2", activebackground="#005fa3")
        btn_lang.place(x=10, y=10)
        self._create_widgets['btn_lang'] = btn_lang

        btn_saveas = tk.Button(dlg, text="💾 " + self._translate("save_as"), bg='#0077b6', fg="white",
                font=button_font, width=14, height=2,
                command=self._save_as, bd=0, relief='flat', cursor="hand2", activebackground="#005fa3")
        btn_saveas.place(relx=1.0, rely=1.0, anchor='se', x=-20, y=-20)
        self._create_widgets['btn_saveas'] = btn_saveas
        self._create_widgets['dlg'] = dlg

        # --- Add tooltips to entry fields in _open_create ---
        def _add_tooltip(widget, text):
            # Simple tooltip implementation
            def on_enter(e):
                self._tip = tk.Toplevel(widget)
                self._tip.wm_overrideredirect(True)
                x = widget.winfo_rootx() + 20
                y = widget.winfo_rooty() + 20
                self._tip.wm_geometry(f"+{x}+{y}")
                label = tk.Label(self._tip, text=text, background="#ffffe0", relief="solid", borderwidth=1, font=("tahoma", "8", "normal"))
                label.pack()
            def on_leave(e):
                if hasattr(self, "_tip"):
                    self._tip.destroy()
            widget.bind("<Enter>", on_enter)
            widget.bind("<Leave>", on_leave)

        # After each Entry, add a tooltip and placeholder
        self.entry_name.insert(0, "e.g. Addis Ketema Iddir")
        _add_tooltip(self.entry_name, "Enter the name of the Iddir (letters and spaces only)")
        self.entry_admin.insert(0, "e.g. Abebe Kebede")
        _add_tooltip(self.entry_admin, "Enter the admin's name (letters and spaces only)")
        self.entry_monthly.insert(0, "e.g. 100")
        _add_tooltip(self.entry_monthly, "Monthly payment amount (number)")
        self.entry_penalty.insert(0, "e.g. 50")
        _add_tooltip(self.entry_penalty, "Penalty amount for late/non-payment (number)")
        self.pin1_entry.insert(0, "")
        _add_tooltip(self.pin1_entry, "Set a numeric PIN for security")
        self.pin2_entry.insert(0, "")
        _add_tooltip(self.pin2_entry, "Confirm your PIN")

        # Warn if closing with unsaved changes
        def on_close():
            if messagebox.askokcancel("Close", "Close without saving?"):
                dlg.destroy()
        dlg.protocol("WM_DELETE_WINDOW", on_close)

    def _on_main_close(self):
        # Confirm before closing main window
        if messagebox.askokcancel("Quit", "Do you want to exit the program? Unsaved changes may be lost."):
            self.destroy()

    def _show_help_dialog(self):
        messagebox.showinfo("Help", "Instructions:\n- Use 'Create' to start a new Iddir file.\n- Use 'Open' to open an existing file.\n- Use the table to manage members and payments.\n- Use the menu to change PIN or view About info.\n- Tooltips are available on input fields.")

    def _show_about_dialog(self):
        messagebox.showinfo(
            "About",
            "Iddir Management System\nVersion 1.0\nDeveloped by:\n- Natnael Belete\n- Biruk Endalkachew\n- Noah Fissiha\n- Yoseph Berie\nFor support, contact: ..."
        )

    def _change_pin_dialog(self):
        dlg = tk.Toplevel(self)
        dlg.title("Change PIN")
        dlg.geometry("350x200")
        dlg.grab_set()
        tk.Label(dlg, text="Old PIN:").pack(pady=(10, 0))
        old_pin_entry = tk.Entry(dlg, show='*')
        old_pin_entry.pack()
        tk.Label(dlg, text="New PIN:").pack(pady=(10, 0))
        new_pin_entry = tk.Entry(dlg, show='*')
        new_pin_entry.pack()
        tk.Label(dlg, text="Confirm New PIN:").pack(pady=(10, 0))
        confirm_pin_entry = tk.Entry(dlg, show='*')
        confirm_pin_entry.pack()
        def do_change_pin():
            old_pin = old_pin_entry.get()
            new_pin = new_pin_entry.get()
            confirm_pin = confirm_pin_entry.get()
            if not os.path.exists(self.pin_file):
                messagebox.showerror("Error", "PIN file missing.")
                return
            with open(self.pin_file) as f:
                iddir_data = json.load(f)
            if old_pin != iddir_data.get("pin"):
                messagebox.showerror("Error", "Old PIN incorrect.")
                return
            if new_pin != confirm_pin or not new_pin:
                messagebox.showerror("Error", "New PINs do not match or are empty.")
                return
            iddir_data["pin"] = new_pin
            with open(self.pin_file, "w") as f:
                json.dump(iddir_data, f)
            messagebox.showinfo("Success", "PIN changed successfully.")
            dlg.destroy()
        tk.Button(dlg, text="Change", command=do_change_pin).pack(pady=15)

    def _switch_language_dialog_inplace(self):
        # Toggle language and rebuild the home screen
        self.language = 'am' if self.language == 'en' else 'en'
        self._build_home()

    # --- Add progress indicator for file operations ---
    def _show_progress(self, msg="Processing..."):
        self._progress_win = tk.Toplevel(self)
        self._progress_win.title("Please wait")
        self._progress_win.geometry("250x80")
        self._progress_win.grab_set()
        tk.Label(self._progress_win, text=msg).pack(pady=20)
        self._progress_win.update()

    def _hide_progress(self):
        if hasattr(self, "_progress_win"):
            self._progress_win.destroy()
            del self._progress_win

    # --- Add keyboard shortcuts for common actions ---
    def _bind_shortcuts(self):
        self.bind_all("<Control-n>", lambda e: self._open_create())
        self.bind_all("<Control-o>", lambda e: self._open_existing())
        self.bind_all("<Control-q>", lambda e: self._on_main_close())
        self.bind_all("<F1>", lambda e: self._show_help_dialog())
        self.bind_all("<Control-l>", lambda e: self._switch_language_dialog_inplace())

# If you have logic for opening files from command-line (double-click), ensure it uses load_from_iddir as well.
if __name__ == "__main__":
    import sys
    # --- Splash screen with logo ---
    def show_splash_and_start():
        splash = tk.Tk()
        splash.overrideredirect(True)
        splash.configure(bg="#ffffff")
        try:
            # Use absolute path for logo if needed
            logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
            logo_img = tk.PhotoImage(file=logo_path)
            w, h = logo_img.width(), logo_img.height()
            splash.geometry(f"{w}x{h+20}+{(splash.winfo_screenwidth()-w)//2}+{(splash.winfo_screenheight()-h)//2}")
            label = tk.Label(splash, image=logo_img, bg="#ffffff")
            label.pack(padx=10, pady=10)
            splash.iconphoto(False, logo_img)
            splash.logo_img = logo_img  # Prevent garbage collection
        except Exception:
            splash.geometry("400x200")
            label = tk.Label(splash, text="Iddir Management", font=("Segoe UI", 24, "bold"), bg="#ffffff")
            label.pack(expand=True)
        splash.after(1800, splash.destroy)  # Show splash for 1.8 seconds
        splash.mainloop()

    show_splash_and_start()  # <-- Ensure splash is shown before app starts

    try:
        app = IddirApp()
        if len(sys.argv) > 1 and sys.argv[1].endswith('.iddir'):
            file_path = sys.argv[1]
            from tkinter import simpledialog
            pin = simpledialog.askstring("PIN", app._translate("enter_pin"), show='*')
            if not pin:
                messagebox.showerror("Error", app._translate("pin_required"))
                sys.exit(1)
            if not os.path.exists(app.pin_file):
                messagebox.showerror("Error", "PIN file missing.")
                sys.exit(1)
            with open(app.pin_file) as f:
                iddir_data = json.load(f)
            if pin != iddir_data.get("pin"):
                messagebox.showerror("Error", app._translate("invalid_pin"))
                sys.exit(1)
            try:
                penalty, columns, data = app.load_from_iddir(file_path)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load file: {e}")
                sys.exit(1)
            app._open_excel_gui(data=data)
            if hasattr(app, '_excel_widgets') and 'penalty_var' in app._excel_widgets:
                app._excel_widgets['penalty_var'].set(str(penalty))
            app.mainloop()
        else:
            app.mainloop()
    except Exception as e:
        import traceback
        traceback.print_exc()
        tk.messagebox.showerror("Fatal Error", f"An error occurred:\n{e}")